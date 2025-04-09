
import os

from PdfPage_ratio import PdfPage
from trace_analysis_ratio import DataFile
import matplotlib.pylab as plt
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np

###get values
def find_nm_files(root_folder):
    nm_paths = []
    
    # Walk through all directories and files in the root_folder
    for folder, _, files in os.walk(root_folder):
        # Check each file in the current directory
        for file in files:

            # Skip files with specific extensions
            if any(ext in file for ext in ['HDF5', 'txt', 'pdf', 'log', 'xlsx']):
                break
            # Construct the full path of the file
            file_path = os.path.join(folder, file)
            normalized_path = os.path.normpath(file_path)
            forward_slash_path = normalized_path.replace("\\", "/")
            nm_paths.append(forward_slash_path)
            print('-', file)

    return nm_paths

def get_datafiles(files, info_df):
    datafiles = []
    for file in files:
        try:
            datafile = DataFile(file, info_df)
            datafiles.append(datafile)
        except Exception as e:
            print(f"Error getting this file : {e}")
    return datafiles    

def get_meta_info(meta_info_directory):
    file_meta_info = open(meta_info_directory, 'r')  
    info_df = pd.read_csv(file_meta_info, header=0, sep=';')
    return info_df

def final_names(files):
    files_id = []
    for file_path in files : 
        file_id = file_path.split('/')[-1].replace('.pxp', '')[2:13]
        files_id.append(file_id) 
        used = set()
        files_id_ = [x for x in files_id if x not in used and (used.add(x) or True)]
    return files_id_

def analyse_datafiles(datafiles, data_list, files_id, debug=False, bis=False, PDF=False, Excel1=False, final_excel=False, barplots=False, STATS=False):
    
    datafiles_ = []
    for datafile in datafiles:

        datafile.calc_values(bis)

        try:
            if bis:
                data_dict = {'Filename': datafile.filename,
                            'Id (A)': datafile.Id_A,
                            'Rm (Ohm)': datafile.Rm,
                            'Ra (Ohm)': datafile.Ra,
                            'Cm (F)': datafile.Cm,
                            'AMPA component (nA)': datafile.amp_resp1,
                            'NMDA component (nA)': datafile.amp_resp2,
                            'NMDA/AMPA ratio': datafile.PPR,
                            'Rise_time 10-90% (ms)': datafile.rise_time,
                            'Decay time 50% (ms)': datafile.decay_time,
                            'Group': datafile.infos['Euthanize method']}
            else:
                data_dict = {'Filename': datafile.filename,
                             'Baseline mean (A)' : datafile.baseline, 
                             #'Baseline STD': datafile.baseline_std,
                             'Id (A)': datafile.Id_A,
                             'Rm (Ohm)': datafile.Rm,
                             'Ra (Ohm)': datafile.Ra,
                             'Cm (F)': datafile.Cm,
                             'Amplitude response 1 (nA)': datafile.amp_resp1,
                             'Amplitude response 2 (nA)': datafile.amp_resp2,
                             'Paired pulse ratio Amp2/Amp1': datafile.PPR,
                             'Rise_time 10-90% (ms)': datafile.rise_time,
                             'Decay time 50% (ms)': datafile.decay_time,
                             'Group': datafile.infos['Euthanize method']}

            data_list.append(data_dict)

            if PDF==True:
                create_pdf(datafile,bis)
            
            if Excel1==True:
                create_excel(data_list, bis)

        except Exception as e:
            print(f"Error analysing this file : {e}")

    #if not bis:
    if final_excel==True:
        create_final_excel(datafiles, files_id)

    if barplots==True:
        create_final_barplots(STATS, bis)

    return 0

###create output
def create_pdf(datafile, bis):
    try:
            pdf = PdfPage(debug=False)
            pdf.fill_PDF(datafile, debug=False, bis=bis)
            #plt.savefig(f'C:/Users/laura.gonzalez/Output_expe/ratio/Ratio_PDFs/{datafile.filename}.pdf')  #plt.savefig(f'C:/Users/LauraGonzalez/Output_expe/Ratio_PDFs/{datafile.filename}.pdf') #laptop
            plt.savefig(f'C:/Users/sofia/Output_expe/In_Vitro/ratio/Ratio_PDFs/{datafile.filename}.pdf')
            print("Individual PDF File saved successfully :", datafile.filename, '\n')

    except Exception as e:
        print(f"Error creating the individual PDF file : {e}")

def create_excel(data_list, bis=False):
    try:
        data_for_excel = pd.DataFrame(data_list)
        #print("data for individual results excel : ", data_for_excel)
        # Create a Pandas Excel writer using openpyxl as the engine  
        if bis:
            #path = 'C:/Users/laura.gonzalez/Output_expe/ratio/ratio_prog_bis.xlsx'
            path = 'C:/Users/sofia/Output_expe/In_Vitro/ratio/ratio_prog_bis.xlsx'
        else:
            #path = 'C:/Users/laura.gonzalez/Output_expe/ratio/ratio_prog.xlsx'
            path = 'C:/Users/sofia/Output_expe/In_Vitro/ratio/ratio_prog.xlsx'
        
        with pd.ExcelWriter(path, engine='openpyxl') as writer: #with pd.ExcelWriter('C:/Users/LauraGonzalez/Output_expe/ratio_prog.xlsx', engine='openpyxl') as writer:
            data_for_excel.to_excel(writer, sheet_name='Data analysis', index=False)
            worksheet = writer.sheets['Data analysis']

            # Adjust column widths for data_for_excel
            for column in data_for_excel:
                column_length = max(data_for_excel[column].astype(str).map(len).max(), len(column))
                col_idx = data_for_excel.columns.get_loc(column)
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = column_length

        print("ratio prog file saved successfully.")

    except Exception as e:
        print(f"ERROR when saving the file to ratio prog : {e}")

def create_final_excel(datafiles, files_id):
    try:
        Ampa1amp, Ampa_rise, Ampa_decay, Ampa2amp  = [],[],[],[]
        Nmda1amp, Nmda_rise, Nmda_decay, Nmda2amp  = [],[],[],[]
        Ratio1, Ratio2 = [], []
        Group,  Group_sh = [], []

        for datafile in datafiles:
            type = datafile.infos['Type']
            Group.append(datafile.infos['Euthanize method'])
            
            if type == 'AMPA' :
                Ampa1amp.append(round(datafile.amp_resp1*(-1)*1000, 2))
                Ampa_rise.append(datafile.rise_time)
                Ampa_decay.append(datafile.decay_time)
                Ampa2amp.append(round(datafile.amp_resp2*(-1)*1000, 2))
      
            elif type == 'NMDA' :
                Nmda1amp.append(round(datafile.amp_resp1*1000, 2))
                Nmda_rise.append(datafile.rise_time)
                Nmda_decay.append(datafile.decay_time)
                Nmda2amp.append(round(datafile.amp_resp2*1000, 2))
            
        data = {'Files_ID'   : files_id,
                "1 AMPA Amplitude (pA)":Ampa1amp , 
                "1 AMPA rise time (10-90%)": Ampa_rise, 
                "1 AMPA decay time (50%)": Ampa_decay, 
                "2 AMPA Amplitude (pA)": Ampa2amp,
                "1 NMDA Amplitude (pA)": Nmda1amp, 
                "1 NMDA rise time (10-90%)": Nmda_rise, 
                "1 NMDA decay time (50%)": Nmda_decay, 
                "2 NMDA Amplitude (pA)": Nmda2amp,
                "1 NMDA/AMPA": None, 
                "2 NMDA/AMPA": None,
                "Group": None}
    
        for j in range(int(len(Group)/2)):
            Group_sh.append(Group[2*j])
        
        for i in range(len(files_id)):
            Ratio1.append(data["1 NMDA Amplitude (pA)"][i]/data["1 AMPA Amplitude (pA)"][i])
            Ratio2.append(data["2 NMDA Amplitude (pA)"][i]/data["2 AMPA Amplitude (pA)"][i])
 
        data["1 NMDA/AMPA"] = Ratio1
        data["2 NMDA/AMPA"] = Ratio2
        data["Group"] = Group_sh

        data_for_excel = pd.DataFrame(data)

        #with pd.ExcelWriter('C:/Users/laura.gonzalez/Output_expe/ratio/final_ratio.xlsx', engine='openpyxl') as writer:
        with pd.ExcelWriter('C:/Users/sofia/Output_expe/In_Vitro/ratio/final_excel.xlsx', engine='openpyxl') as writer:
    
            data_for_excel.to_excel(writer, sheet_name='Data analysis', index=False)
            # Access the workbook and the sheets
            workbook  = writer.book
            worksheet = writer.sheets['Data analysis']

            # Adjust column widths for data_mem_for_excel
            for column in data_for_excel:
                column_length = max(data_for_excel[column].astype(str).map(len).max(), len(column))
                col_idx = data_for_excel.columns.get_loc(column)
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = column_length
    
        print("Final Excel file saved successfully.")


    except Exception as e:
        print(f"ERROR when saving the file to final excel : {e}")

def create_final_barplots(STATS, bis=False):
    try:
        
        final_excel_path = 'C:/Users/sofia/Output_expe/In_Vitro/ratio/final_excel.xlsx' #compare with IGOR results

        metrics = ["1 AMPA Amplitude (pA)", 
                   "1 AMPA rise time (10-90%)", 
                   "1 AMPA decay time (50%)", 
                   "2 AMPA Amplitude (pA)",
                   "1 NMDA Amplitude (pA)", 
                   "1 NMDA rise time (10-90%)", 
                   "1 NMDA decay time (50%)", 
                   "2 NMDA Amplitude (pA)",
                   "1 NMDA/AMPA", 
                   "2 NMDA/AMPA"]  
        
        pdf = PdfPage(debug=False)
        pdf.fill_PDF_barplots(final_excel_path, metrics, STATS)
        plt.savefig(f'C:/Users/sofia/Output_expe/In_Vitro/ratio/Ratio_PDFs/auto_barplots.pdf')
        print("Final barolots PDF file saved successfully.")

    except Exception as e:
        print(f"Error saving the barplots : {e}")
    return 0

###### MAIN ######################################################

if __name__=='__main__':
    debug = True
    bis   = False #if true, 2nd version of AMPA NMDA ratio is calculated

    if bis: 
        #files_directory = 'C:/Users/laura.gonzalez/DATA/Ratio_experiment/RAW_DATA_AMPA_NMDA_RATIO-bis'
        #meta_info_directory = 'C:/Users/laura.gonzalez/DATA/Ratio_experiment/Files-bis.csv'
        files_directory = 'C:/Users/sofia/DATA/In_Vitro_experiments/Ratio_experiment/RAW_DATA_AMPA_NMDA_RATIO-bis'
        meta_info_directory = 'C:/Users/sofia/DATA/In_Vitro_experiments/Ratio_experiment/Files-bis.csv'

    else:
        #files_directory = 'C:/Users/laura.gonzalez/DATA/Ratio_experiment/RAW_DATA_AMPA_NMDA_RATIO'  #PC   #files_directory = 'C:/Users/LauraGonzalez/DATA/Ratio_experiment/RAW_DATA_AMPA_NMDA_RATIO-q' #laptop
        #meta_info_directory = 'C:/Users/laura.gonzalez/DATA/Ratio_experiment/Files.csv'
        files_directory = 'C:/Users/sofia/DATA/In_Vitro_experiments/Ratio_experiment/RAW_DATA_AMPA_NMDA_RATIO-q'  #PC   #files_directory = 'C:/Users/LauraGonzalez/DATA/Ratio_experiment/RAW_DATA_AMPA_NMDA_RATIO-q' #laptop
        meta_info_directory = 'C:/Users/sofia/DATA/In_Vitro_experiments/Ratio_experiment/Files.csv'
        
    files = find_nm_files(files_directory)
    info_df = get_meta_info(meta_info_directory)
    
    datafiles = get_datafiles(files, info_df)
    data_list = []
    files_id = final_names(files)

    #Choose as True what you want to plot
    analyse_datafiles(datafiles, 
                      data_list, 
                      files_id, 
                      debug=debug,
                      bis=bis, 
                      PDF=False, 
                      Excel1=True, 
                      final_excel=True, 
                      barplots=False, 
                      STATS=False)