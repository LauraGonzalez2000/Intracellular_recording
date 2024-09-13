
import os

from PdfPage_ratio import PdfPage
from trace_analysis_ratio import DataFile
import matplotlib.pylab as plt
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np

files_directory = 'C:/Users/laura.gonzalez/DATA/Ratio_experiment/RAW_DATA_AMPA_NMDA_RATIO'  #PC   #files_directory = 'C:/Users/LauraGonzalez/DATA/Ratio_experiment/RAW_DATA_AMPA_NMDA_RATIO-q' #laptop

def find_nm_files(root_folder):
    nm_paths = []
    
    # Walk through all directories and files in the root_folder
    for folder, _, files in os.walk(root_folder):
        # Check each file in the current directory
        for file in files:

            # Skip files with specific extensions
            if any(ext in file for ext in ['HDF5', 'txt', 'pdf', 'log', 'xlsx']):
                break
            # Construct the full path of the file
            file_path = os.path.join(folder, file)
            normalized_path = os.path.normpath(file_path)
            forward_slash_path = normalized_path.replace("\\", "/")
            nm_paths.append(forward_slash_path)
            #print('-', file)

    return nm_paths

def plot_barplots(file_path, metrics):
    # Read the Excel file into a DataFrame
    IGOR_results_df = pd.read_excel(file_path, header=0)
    #print(IGOR_results_df)
    
    # Group labels
    labels = ['Ketaxyla', 'xyla eutha', 'keta xyla eutha']
    group_queries = ["Group=='ketaxyla'", "Group=='xyla euthasol'", "Group=='keta xyla euthasol'"]
    colors = ['blue', 'green', 'orange']
    
    # Number of plots based on the number of metrics
    num_metrics = len(metrics)
    
    # Set up the figure and subplots
    fig, axes = plt.subplots(3, 4, figsize=(14, 16))
    axes = axes.flatten()
    
    #plt.subplots(1, num_metrics, figsize=(10 * num_metrics, 6), sharey=False)
    
    # Ensure axes is always a list (in case there's only one metric)
    if num_metrics == 1:
        axes = [axes]
    
    # Loop through each metric and plot
    for idx, metric in enumerate(metrics):
        means = []
        sems = []
        individual_data = []
        
        # Collect data for each group
        for query in group_queries:
            df_temp = IGOR_results_df.query(query)[metric]
            means.append(df_temp.mean())
            sems.append(df_temp.sem())
            individual_data.append(df_temp)
        
        # Bar plot with error bars
        bar_positions = np.arange(len(labels))
        axes[idx].bar(bar_positions, means, yerr=sems, color=colors, width=0.6, capsize=5, alpha=0.6, label='Mean ± SEM')
        
        # Plot individual data points
        for i, df in enumerate(individual_data):
            axes[idx].scatter(np.full(df.shape, bar_positions[i]), df, color='black', zorder=5)
        
        axes[idx].spines['top'].set_visible(False)
        axes[idx].spines['right'].set_visible(False)
        
        # Add title and labels to each subplot
        axes[idx].set_title(f'{metric}')
        axes[idx].set_xticks(bar_positions)
        axes[idx].set_xticklabels(labels)
        #axes[idx].set_ylabel('Amplitude (pA)')

    # Delete any remaining unused subplots (if any)
    for idx in range(num_metrics, len(axes)):
        fig.delaxes(axes[idx])
    
    # Adjust layout to prevent overlap
    plt.tight_layout()
    #plt.show()

def get_datafiles(files):
    datafiles = []
    for file in files:
        try:
            datafile = DataFile(file)
            datafiles.append(datafile)
        except Exception as e:
            print(f"Error getting this file : {e}")
    return datafiles

###### MAIN ######################################################

if __name__=='__main__':
    
    files = find_nm_files(files_directory)
    data_list = []
    
    datafiles = get_datafiles(files)
    #print("len datafiles ",len(datafiles))

    #PDF creation per file
    for datafile in datafiles:
        try:
            pdf = PdfPage(debug=False)
            pdf.fill_PDF(datafile, debug=False)
            plt.savefig(f'C:/Users/laura.gonzalez/Output_expe/Ratio_PDFs/{datafile.filename}.pdf')  #plt.savefig(f'C:/Users/LauraGonzalez/Output_expe/Ratio_PDFs/{datafile.filename}.pdf') #laptop
            print("Individual PDF File saved successfully :", datafile.filename, '\n')

            data_dict = {'Filename': datafile.filename,
                            'Id (A)': datafile.Id_A,
                            'Rm (Ohm)': datafile.Rm,
                            'Ra (Ohm)': datafile.Ra,
                            'Cm (F)': datafile.Cm,
                            'Peak type' : datafile.type,
                            'Amplitude response 1 (nA)': datafile.amp_resp1,
                            'Amplitude response 2 (nA)': datafile.amp_resp2,
                            'Paired pulse ratio Amp2/Amp1': datafile.PPR,
                            'Rise_time 10-90% (ms)': datafile.rise_time,
                            'Decay time 50% (ms)': datafile.decay_time,
                            'Group': datafile.infos['Euthanize method']}
            data_list.append(data_dict)
            
        except Exception as e:
            print(f"Error creating the individual PDF file : {e}")

    #Excel creation
    try:
        data_for_excel = pd.DataFrame(data_list)
        print("data for individual results excel : ", data_for_excel)
        # Create a Pandas Excel writer using openpyxl as the engine  
        with pd.ExcelWriter('C:/Users/laura.gonzalez/Output_expe/ratio_prog.xlsx', engine='openpyxl') as writer: #with pd.ExcelWriter('C:/Users/LauraGonzalez/Output_expe/ratio_prog.xlsx', engine='openpyxl') as writer:
            data_for_excel.to_excel(writer, sheet_name='Data analysis', index=False)
            # Access the workbook and the sheets
            workbook  = writer.book
            worksheet = writer.sheets['Data analysis']

            # Adjust column widths for data_for_excel
            for column in data_for_excel:
                column_length = max(data_for_excel[column].astype(str).map(len).max(), len(column))
                col_idx = data_for_excel.columns.get_loc(column)
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = column_length

        print("Excel 1 file saved successfully.")

    except Exception as e:
        print(f"ERROR when saving the file to excel1 : {e}")


    #Final excel creation
    files_id = []
    try:
        file_paths = find_nm_files(files_directory)
        for file_path in file_paths : 
            file_id = file_path.split('/')[-1].replace('.pxp', '')[2:13]
            files_id.append(file_id) 

        used = set()
        files_id_ = [x for x in files_id if x not in used and (used.add(x) or True)]
        #print("final files ID: ", files_id_)

        Ampa1amp = []
        Ampa_rise = []
        Ampa_decay = []
        Ampa2amp = []
        Nmda1amp = []
        Nmda_rise = []
        Nmda_decay = []
        Nmda2amp = []
        Ratio1 = []
        Ratio2 = []
        Group = []
        Group_sh = []

        for datafile in datafiles:
            id = datafile.filename[2:13]
            type = datafile.infos['Type']
            #print(id)
            #print(datafile.infos['Type'])

            Group.append(datafile.infos['Euthanize method'])

            if type == 'AMPA' :
                Ampa1amp.append(round(datafile.amp_resp1*(-1)*1000, 2))
                Ampa_rise.append(datafile.rise_time)
                Ampa_decay.append(datafile.decay_time)
                Ampa2amp.append(round(datafile.amp_resp2*(-1)*1000, 2))
                

            elif type == 'NMDA' :
                Nmda1amp.append(round(datafile.amp_resp1*1000, 2))
                Nmda_rise.append(datafile.rise_time)
                Nmda_decay.append(datafile.decay_time)
                Nmda2amp.append(round(datafile.amp_resp2*1000, 2))

        #print("aa",Nmda2amp)
        #print("bb",Group)
            
        data = {'Files_ID'   : files_id_,
                "1 AMPA Amplitude (pA)":Ampa1amp , 
                "1 AMPA rise time (10-90%)": Ampa_rise, 
                "1 AMPA decay time (50%)": Ampa_decay, 
                "2 AMPA Amplitude (pA)": Ampa2amp,
                "1 NMDA Amplitude (pA)": Nmda1amp, 
                "1 NMDA rise time (10-90%)": Nmda_rise, 
                "1 NMDA decay time (50%)": Nmda_decay, 
                "2 NMDA Amplitude (pA)": Nmda2amp,
                "1 NMDA/AMPA": None, 
                "2 NMDA/AMPA": None,
                "Group": None}
        
        for j in range(len(files_id_)):
            Group_sh.append(Group[j])
            j+=2
        
        for i in range(len(files_id_)):
            Ratio1.append(data["1 NMDA Amplitude (pA)"][i]/data["1 AMPA Amplitude (pA)"][i])
            Ratio2.append(data["2 NMDA Amplitude (pA)"][i]/data["2 AMPA Amplitude (pA)"][i])

        data["1 NMDA/AMPA"] = Ratio1
        data["2 NMDA/AMPA"] = Ratio2
        data["Group"] = Group_sh

        
        data_for_excel = pd.DataFrame(data)
        #print(data_for_excel)

        with pd.ExcelWriter('C:/Users/laura.gonzalez/Output_expe/final_ratio.xlsx', engine='openpyxl') as writer:
            
            data_for_excel.to_excel(writer, sheet_name='Data analysis', index=False)
            # Access the workbook and the sheets
            workbook  = writer.book
            worksheet = writer.sheets['Data analysis']

            # Adjust column widths for data_mem_for_excel
            for column in data_for_excel:
                column_length = max(data_for_excel[column].astype(str).map(len).max(), len(column))
                col_idx = data_for_excel.columns.get_loc(column)
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = column_length
        print("Excel 2 file saved successfully.")
    except Exception as e:
        print(f"ERROR when saving the file to excel1 : {e}")


    #Final Barlots
    manual_results_path = 'C:/Users/laura.gonzalez/Output_expe/ratio_results_.xlsx'
    automatic_results_path = 'C:/Users/laura.gonzalez/Output_expe/final_ratio.xlsx' #compare with IGOR results
    metrics = ["1 AMPA Amplitude (pA)", "1 AMPA rise time (10-90%)", "1 AMPA decay time (50%)", "2 AMPA Amplitude (pA)",
            "1 NMDA Amplitude (pA)", "1 NMDA rise time (10-90%)", "1 NMDA decay time (50%)", "2 NMDA Amplitude (pA)",
            "1 NMDA/AMPA", "2 NMDA/AMPA"]  
    #plot_barplots(manual_results_path, metrics)
    #plt.savefig(f'C:/Users/laura.gonzalez/Output_expe/Ratio_PDFs/manual_barplots.pdf')
    
    plot_barplots(automatic_results_path, metrics) #PC  #plot_barplots('C:/Users/LauraGonzalez/Output_expe/ratio_results_.xlsx', metrics) #laptop
    plt.savefig(f'C:/Users/laura.gonzalez/Output_expe/Ratio_PDFs/auto_barplots.pdf')


