
import os

from PdfPage import PdfPage
from trace_analysis import DataFile
from igor2.packed import load as loadpxp
import pprint
import matplotlib.pylab as plt
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
#from openpyxl import load_workbook

def find_nm_files(root_folder):
    nm_paths = []
    
    # Walk through all directories and files in the root_folder
    for folder, _, files in os.walk(root_folder):
        # Check each file in the current directory
        for file in files:

            # Skip files with specific extensions
            if any(ext in file for ext in ['HDF5', 'txt', 'pdf', 'log', 'xlsx']):
                break
            # Construct the full path of the file
            file_path = os.path.join(folder, file)
            normalized_path = os.path.normpath(file_path)
            forward_slash_path = normalized_path.replace("\\", "/")
            nm_paths.append(forward_slash_path)
            #print('-', file)

    return nm_paths


###### MAIN ######################################################

#files = find_nm_files('C:/Users/laura.gonzalez/DATA/DATA_TO_ANALYSE')
files = find_nm_files('D:\Internship_Rebola_ICM\RAW_DATA_TO_ANALYSE')

data_mem_list = []
data_resp_list = []

#PDF creation per file
for file in files:
    try:
        datafile = DataFile(file)
        pdf = PdfPage(debug=False)
        pdf.fill_PDF(datafile, debug=False)
        plt.savefig(f'C:/Users/laura.gonzalez/DATA/PDFs/{datafile.filename}.pdf')
        print("File saved successfully :", file, '\n')

        
        data_mem_dict = {'Filename': datafile.filename,
                        'Id (A)': datafile.Id_A,
                        'Rm (Ohm)': datafile.Rm,
                        'Ra (Ohm)': datafile.Ra,
                        'Cm (F)': datafile.Cm}
        data_mem_list.append(data_mem_dict)

        data_resp_dict = {'Filename': datafile.filename,
                          'Peak type' : datafile.type,
                          'Amplitude response 1 (nA)': datafile.amp_resp1,
                          'Amplitude response 2 (nA)': datafile.amp_resp2,
                          'Paired pulse ratio Amp2/Amp1': datafile.PPR,
                          'Rise_time 10-90% (ms)': datafile.rise_time,
                          'Decay time 50% (ms)': datafile.decay_time}
        data_resp_list.append(data_resp_dict)
    except Exception as e:
        print(f"Error analysing this file : {e}")

#Excel creation

try:
    data_mem_for_excel = pd.DataFrame(data_mem_list)
    data_resp_for_excel = pd.DataFrame(data_resp_list)

    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter('C:/Users/laura.gonzalez/DATA/output.xlsx', engine='openpyxl') as writer:
        data_mem_for_excel.to_excel(writer, sheet_name='Membrane characteristics', index=False)
        data_resp_for_excel.to_excel(writer, sheet_name='Response', index=False)

        # Access the workbook and the sheets
        workbook  = writer.book
        worksheet_mem = writer.sheets['Membrane characteristics']
        worksheet_resp = writer.sheets['Response']

        # Adjust column widths for data_mem_for_excel
        for column in data_mem_for_excel:
            column_length = max(data_mem_for_excel[column].astype(str).map(len).max(), len(column))
            col_idx = data_mem_for_excel.columns.get_loc(column)
            worksheet_mem.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = column_length

        # Adjust column widths for data_resp_for_excel
        for column in data_resp_for_excel:
            column_length = max(data_resp_for_excel[column].astype(str).map(len).max(), len(column))
            col_idx = data_resp_for_excel.columns.get_loc(column)
            worksheet_resp.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = column_length

    print("Excel file saved successfully.")
except Exception as e:
    print(f"ERROR when saving the file to excel: {e}")


#analyse each file to plot

#Execute only if the Python script is being executed as the main program. 
'''
if __name__=='__main__':
    
    #filename = os.path.join(os.path.expanduser('~'), 'DATA', 'Dataset1', 'nm12Jun2024c0_000_AMPA.pdf')
    datafile = DataFile('D:/Internship_Rebola_ICM/DATA_TO_ANALYSE/nm28May2024c1/nm28May2024c1_001.pxp')
    #datafile = DataFile('C:/Users/laura.gonzalez/DATA/RAW_DATA/model_cell/nm24Jun2024c0_000.pxp')
    pdf = PdfPage(debug=True)
    pdf.fill_PDF(datafile, debug=True)
    plt.show()
'''
